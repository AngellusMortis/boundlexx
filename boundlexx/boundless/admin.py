from datetime import timedelta
from typing import Sequence, Type

from django.conf import settings
from django.contrib import admin
from django.contrib.admin.options import InlineModelAdmin
from django.http import HttpResponse
from django.templatetags.tz import localtime
from django.urls import reverse
from django.utils import timezone
from django.utils.formats import localize
from django.utils.safestring import mark_safe
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook

from boundlexx.boundless.models import (
    Color,
    ColorValue,
    Item,
    ItemBuyRank,
    ItemRequestBasketPrice,
    ItemSellRank,
    ItemShopStandPrice,
    LeaderboardRecord,
    LocalizedName,
    Metal,
    Recipe,
    RecipeGroup,
    RecipeLevel,
    ResourceCount,
    Skill,
    SkillGroup,
    Subtitle,
    World,
    WorldBlockColor,
    WorldCreatureColor,
    WorldPoll,
    WorldPollResult,
)

TIMESERIES_CUTOFF = 7


class LocalizationInline(admin.TabularInline):
    model = LocalizedName
    readonly_fields = [
        "lang",
        "name",
    ]
    can_delete = False
    max_num = 0


class ItemPriceInline(admin.TabularInline):
    fk_name = "item"
    readonly_fields = [
        "time",
        "world",
        "location_x",
        "location_y",
        "location_z",
        "price",
        "item_count",
        "beacon_name",
        "guild_tag",
        "shop_activity",
    ]
    can_delete = False
    max_num = 0

    def get_queryset(self, request):
        cutoff = timezone.now() - timedelta(days=TIMESERIES_CUTOFF)
        return super().get_queryset(request).filter(active=True, time__gt=cutoff)


class ItemRankInline(admin.TabularInline):
    readonly_fields = [
        "world",
        "rank",
        "last_update",
        "state_hash",
        "next_update",
    ]
    can_delete = False
    max_num = 0

    def next_update(self, obj):
        return localize(localtime(obj.next_update))


class ItemBuyRankInline(ItemRankInline):
    model = ItemBuyRank


class ItemSellRankInline(ItemRankInline):
    model = ItemSellRank


class ItemRequestBasketPriceInline(ItemPriceInline):
    model = ItemRequestBasketPrice


class ItemShopStandPriceInline(ItemPriceInline):
    model = ItemShopStandPrice


class ColorValueInline(admin.TabularInline):
    model = ColorValue

    readonly_fields = [
        "color_type",
        "shade",
        "base",
        "hlight",
        "rgb_color",
    ]
    can_delete = False
    max_num = 0


class ItemResourceCountInline(admin.TabularInline):
    model = ResourceCount
    can_delete = False
    max_num = 0
    raw_id_fields = ["world_poll"]

    def world(self, obj):
        return obj.world_poll.world

    fields = [
        "world",
        "count",
    ]
    readonly_fields = [
        "world",
        "count",
    ]

    def get_queryset(self, request):
        return (
            super()
            .get_queryset(request)
            .filter(world_poll__active=True, world_poll__world__active=True)
            .select_related("world_poll__world")
            .order_by("-count")
        )


class GameObjAdmin(admin.ModelAdmin):
    list_display = ["game_id", "default_name", "active"]
    search_fields = ["game_id", "localizedname__name"]
    readonly_fields = ["game_id"]

    inlines: Sequence[Type[InlineModelAdmin]] = [LocalizationInline]

    def has_delete_permission(self, request, obj=None):
        return False

    def get_queryset(self, request):
        return super().get_queryset(request).prefetch_related("localizedname_set")


@admin.register(Subtitle)
class SubtitleAdmin(GameObjAdmin):
    pass


@admin.register(Color)
class ColorAdmin(GameObjAdmin):
    list_display = [
        "game_id",
        "default_name",
        "base_color",
        "gleam_color",
        "active",
    ]
    readonly_fields = [
        "game_id",
        "base_color",
        "gleam_color",
    ]

    inlines = [ColorValueInline]

    def get_queryset(self, request):
        return (
            super()
            .get_queryset(request)
            .prefetch_related("localizedname_set", "colorvalue_set")
        )


@admin.register(Metal)
class MetalAdmin(GameObjAdmin):
    pass


@admin.register(Item)
class ItemAdmin(GameObjAdmin):
    def is_resource(self, obj):
        return obj.is_resource

    def has_colors(self, obj):
        return obj.has_colors

    is_resource.boolean = True  # type: ignore
    has_colors.boolean = True  # type: ignore

    list_display = [
        "game_id",
        "default_name",
        "is_resource",
        "has_colors",
        "active",
    ]
    readonly_fields = [
        "game_id",
        "string_id",
        "is_resource",
        "has_colors",
    ]
    raw_id_fields = ["item_subtitle"]
    search_fields = ["game_id", "string_id", "localizedname__name"]

    def get_inlines(self, request, obj):
        inlines = [
            LocalizationInline,
            ItemBuyRankInline,
            ItemRequestBasketPriceInline,
            ItemSellRankInline,
            ItemShopStandPriceInline,
        ]

        if obj.game_id in settings.BOUNDLESS_WORLD_POLL_RESOURCE_MAPPING:
            inlines.insert(1, ItemResourceCountInline)

        return inlines

    def get_queryset(self, request):
        return (
            super()
            .get_queryset(request)
            .prefetch_related("localizedname_set", "worldblockcolor_set")
        )


class ItemPriceAdmin(admin.ModelAdmin):
    raw_id_fields = ["item"]
    list_display = [
        "item",
        "item_count",
        "world",
        "location",
        "price",
        "beacon_name",
        "active",
    ]
    readonly_fields = [
        "time",
        "world",
        "location_x",
        "location_y",
        "location_z",
        "price",
        "item_count",
        "beacon_name",
        "guild_tag",
        "shop_activity",
    ]
    search_fields = [
        "item__string_id",
        "item__localizedname__name",
        "world__display_name",
        "beacon_name",
    ]


@admin.register(ItemRequestBasketPrice)
class ItemRequestBasketPriceAdmin(ItemPriceAdmin):
    pass


@admin.register(ItemShopStandPrice)
class ItemShopStandPriceAdmin(ItemPriceAdmin):
    pass


class WorldPollInline(admin.TabularInline):
    model = WorldPoll

    show_change_link = True
    fields = ["active", "time"]
    readonly_fields = ["time"]
    can_delete = False
    max_num = 0

    def get_queryset(self, request):
        cutoff = timezone.now() - timedelta(days=TIMESERIES_CUTOFF)
        return super().get_queryset(request).filter(time__gt=cutoff).order_by("-time")


class WorldBlockColorInline(admin.TabularInline):
    fk_name = "world"
    model = WorldBlockColor

    fields = [
        "item",
        "color",
    ]
    readonly_fields = [
        "item",
        "color",
    ]
    can_delete = False
    max_num = 0


class WorldCreatureColorInline(admin.TabularInline):
    model = WorldCreatureColor

    fields = ["creature_type", "color_data"]
    readonly_fields = ["creature_type", "color_data"]
    can_delete = False
    max_num = 0


@admin.register(World)
class WorldAdmin(admin.ModelAdmin):
    def is_perm(self, obj):
        return obj.is_perm

    def is_exo(self, obj):
        return obj.is_exo

    def is_sovereign(self, obj):
        return obj.is_sovereign

    is_perm.boolean = True  # type: ignore
    is_exo.boolean = True  # type: ignore
    is_sovereign.boolean = True  # type: ignore

    list_display = [
        "id",
        "display_name",
        "image",
        "forum_url",
        "region",
        "tier",
        "world_type",
        "protection",
        "active",
        "is_perm",
        "is_exo",
        "is_creative",
        "is_sovereign",
        "is_public",
        "is_locked",
    ]
    fields = [
        "id",
        "active",
        "is_perm",
        "is_creative",
        "is_sovereign",
        "is_public",
        "is_locked",
        "name",
        "display_name",
        "region",
        "tier",
        "size",
        "number_of_regions",
        "world_type",
        "address",
        "ip_address",
        "api_url",
        "planets_url",
        "chunks_url",
        "time_offset",
        "websocket_url",
        "atmosphere_color",
        "water_color",
        "start",
        "end",
        "surface_liquid",
        "core_liquid",
        "forum_id",
    ]
    readonly_fields = [
        "id",
        "is_perm",
        "is_creative",
        "is_sovereign",
        "is_public",
        "is_locked",
        "name",
        "display_name",
        "region",
        "tier",
        "size",
        "number_of_regions",
        "world_type",
        "address",
        "ip_address",
        "api_url",
        "planets_url",
        "chunks_url",
        "time_offset",
        "websocket_url",
        "atmosphere_color",
        "water_color",
        "start",
        "end",
        "surface_liquid",
        "core_liquid",
        "forum_id",
    ]
    search_fields = ["display_name", "world_type"]

    inlines = [
        WorldPollInline,
        WorldBlockColorInline,
        WorldCreatureColorInline,
    ]

    def get_queryset(self, request):
        return (
            super()
            .get_queryset(request)
            .prefetch_related(
                "worldpoll_set",
                "worldblockcolor_set",
                "worldblockcolor_set__item",
                "worldblockcolor_set__color",
                "worldblockcolor_set__color__localizedname_set",
                "worldcreaturecolor_set",
            )
        )


class WorldPollResultInline(admin.StackedInline):
    model = WorldPollResult

    readonly_fields = [
        "player_count",
        "beacon_count",
        "plot_count",
        "total_prestige",
    ]
    can_delete = False
    max_num = 0


class ResourceCountInline(admin.TabularInline):
    model = ResourceCount

    def is_embedded(self, obj):
        return obj.is_embedded

    is_embedded.boolean = True  # type: ignore

    fields = [
        "item",
        "is_embedded",
        "percentage",
        "count",
    ]
    readonly_fields = [
        "item",
        "is_embedded",
        "percentage",
        "count",
    ]
    can_delete = False
    max_num = 0

    def get_queryset(self, request):
        return super().get_queryset(request).order_by("-count")


class LeaderboardRecordInline(admin.TabularInline):
    model = LeaderboardRecord

    fields = [
        "world_rank",
        "guild_tag",
        "mayor_name",
        "name",
        "prestige",
    ]
    readonly_fields = [
        "world_rank",
        "guild_tag",
        "mayor_name",
        "name",
        "prestige",
    ]
    can_delete = False
    max_num = 0

    def get_queryset(self, request):
        return super().get_queryset(request).order_by("world_rank")


@admin.register(WorldPoll)
class WorldPollAdmin(admin.ModelAdmin):
    list_display = ["world", "time", "active"]

    fields = ["active", "world", "time"]
    readonly_fields = ["time"]

    inlines = [
        WorldPollResultInline,
        LeaderboardRecordInline,
        ResourceCountInline,
    ]

    def get_queryset(self, request):
        return (
            super()
            .get_queryset(request)
            .select_related("world")
            .prefetch_related(
                "worldpollresult_set",
                "leaderboardrecord_set",
                "resourcecount_set",
                "resourcecount_set__item",
            )
        )


@admin.register(SkillGroup)
class SkillGroupAdmin(admin.ModelAdmin):
    list_display = ["name", "skill_type", "unlock_level"]


@admin.register(Skill)
class SkillAdmin(admin.ModelAdmin):
    list_display = ["name", "group", "number_unlocks", "cost"]

    def get_queryset(self, request):
        return super().get_queryset(request).select_related("group")


@admin.register(RecipeGroup)
class RecipeGroupAdmin(admin.ModelAdmin):
    list_display = ["name"]

    readonly_fields = ["name", "display_name", "members"]


class RecipeInputInline(admin.TabularInline):
    model = RecipeLevel.inputs.through

    def input_group(self, obj):
        return obj.recipeinput.group or "-"

    input_group.short_description = "group"  # type: ignore

    def input_item(self, obj):
        return obj.recipeinput.item or "-"

    input_item.short_description = "item"  # type: ignore

    def input_count(self, obj):
        return obj.recipeinput.count

    input_count.short_description = "count"  # type: ignore

    fields = [
        "input_group",
        "input_item",
        "input_count",
    ]
    readonly_fields = [
        "input_group",
        "input_item",
        "input_count",
    ]
    can_delete = False
    max_num = 0


@admin.register(RecipeLevel)
class RecipeLevelAdmin(admin.ModelAdmin):
    readonly_fields = [
        "level",
        "wear",
        "spark",
        "duration",
        "output_quantity",
    ]

    exclude = ["inputs"]

    inlines = [
        RecipeInputInline,
    ]

    def get_queryset(self, request):
        return super().get_queryset(request).prefetch_related("inputs")


class RecipeRequirementsInline(admin.TabularInline):
    model = Recipe.requirements.through

    def requirements_skill(self, obj):
        return obj.reciperequirement.skill

    requirements_skill.short_description = "skill"  # type: ignore

    def requirements_level(self, obj):
        return obj.reciperequirement.level

    requirements_level.short_description = "level"  # type: ignore

    fields = [
        "requirements_skill",
        "requirements_level",
    ]
    readonly_fields = [
        "requirements_skill",
        "requirements_level",
    ]
    can_delete = False
    max_num = 0


class RecipeLevelInline(admin.TabularInline):
    model = Recipe.levels.through

    def level(self, obj):
        return obj.recipelevel.get_level_display()

    level.short_description = "level"  # type: ignore

    def link(self, obj):
        link = reverse("admin:boundless_recipelevel_change", args=(obj.recipelevel.id,))

        return mark_safe(f'<a href="{link}">View</a>')  # nosec

    link.short_description = "link"  # type: ignore

    fields = [
        "level",
        "link",
    ]
    readonly_fields = [
        "level",
        "link",
    ]
    can_delete = False
    max_num = 0


@admin.register(Recipe)
class RecipeAdmin(GameObjAdmin):
    list_display = [
        "output",
        "group_name",
        "machine",
        "craft_xp",
        "can_hand_craft",
    ]

    readonly_fields = [
        "game_id",
        "machine",
        "heat",
        "craft_xp",
        "output",
        "can_hand_craft",
        "machine_level",
        "power",
        "group_name",
        "knowledge_unlock_level",
        "tints",
    ]

    exclude = ["requirements", "levels"]

    inlines = [RecipeRequirementsInline, RecipeLevelInline]
    actions = ["export_to_xlsx"]

    def _get_levels(self, recipe):
        every = [
            recipe.output.english,
            recipe.can_hand_craft,
            recipe.machine,
            recipe.heat if recipe.machine == "FURNACE" else recipe.power,
            recipe.group_name,
            ",".join(i.english for i in recipe.tints.all()),
            ",".join(f"{r.skill.name} Lvl{r.level}" for r in recipe.requirements.all()),
            recipe.get_required_event_display(),
            recipe.required_backer_tier,
        ]

        single_row = every[:]
        bulk_row = every[:]
        mass_row = every[:]

        single_only = False

        for level in recipe.levels.all():
            if recipe.machine == "FURNACE":
                if level.level == 0:
                    extra_row = [
                        level.wear,
                        level.duration,
                        level.output_quantity,
                        recipe.craft_xp * level.output_quantity,
                    ]
                else:
                    continue
            else:
                extra_row = [
                    level.wear,
                    level.spark,
                    level.duration,
                    level.output_quantity,
                    recipe.craft_xp * level.output_quantity,
                ]

            if level.inputs.count() == 0:
                single_only = True

            for input_item in level.inputs.all():
                if input_item.group is None:
                    extra_row += [input_item.item.english, input_item.count]
                else:
                    extra_row += [input_item.group.name, input_item.count]

            if level.level == 0:
                single_row += extra_row
            elif level.level == 1:
                bulk_row += extra_row
            else:
                mass_row += extra_row

        return single_only, [single_row, bulk_row, mass_row]

    def export_to_xlsx(self, request, queryset):
        queryset = queryset.prefetch_related(
            "tints",
            "tints__localizedname_set",
            "requirements",
            "requirements__skill",
            "levels",
            "levels__inputs",
            "levels__inputs__group",
            "levels__inputs__item",
            "levels__inputs__item__localizedname_set",
        )
        groups = RecipeGroup.objects.all().prefetch_related(
            "members",
            "members__localizedname_set",
        )

        workbook = Workbook()
        workbook.active.title = "Single"
        workbook.create_sheet("Bulk")
        workbook.create_sheet("Mass")
        workbook.create_sheet("Furnace")
        workbook.create_sheet("Groups")

        group_sheet = workbook["Groups"]
        group_sheet.append(["Name", "Members"])

        for group in groups:
            group_sheet.append([group.name] + [i.english for i in group.members.all()])

        single = workbook["Single"]
        bulk = workbook["Bulk"]
        mass = workbook["Mass"]
        furnace = workbook["Furnace"]

        headers = [
            "Output",
            "Hand Craft?",
            "Machine",
            "Power",
            "Group",
            "Tint From",
            "Requirements",
            "Required Event",
            "Required Backer",
            "Machine Wear",
            "Spark",
            "Duration (s)",
            "Output Quantity",
            "XP",
            "Inputs",
        ]
        single.append(headers)
        bulk.append(headers)
        mass.append(headers)
        furnace.append(
            [
                "Output",
                "Hand Craft?",
                "Machine",
                "Heat",
                "Group",
                "Tint From",
                "Requirements",
                "Required Event",
                "Required Backer",
                "Machine Wear",
                "Duration (s)",
                "Output Quantity",
                "XP",
                "Inputs",
            ]
        )
        for recipe in queryset:
            single_only, rows = self._get_levels(recipe)

            if recipe.machine == "FURNACE":
                furnace.append(rows[0])
                continue

            if rows[0][12] == rows[1][12]:
                single_only = True

            single.append(rows[0])
            if not single_only:
                bulk.append(rows[1])
                mass.append(rows[2])

        return HttpResponse(
            save_virtual_workbook(workbook),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # noqa: E501
        )

    export_to_xlsx.short_description = "Export to Excel"  # type: ignore
