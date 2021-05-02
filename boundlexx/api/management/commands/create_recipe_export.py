import djclick as click
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook

from boundlexx.api.tasks import purge_static_cache
from boundlexx.api.utils import create_export_file, set_column_widths
from boundlexx.boundless.models import Recipe, RecipeGroup

FILENAME = "recipe_export"
DESCRIPTION = """All recipes in the game exported in a Excel spreadsheet.
"""

HEADERS = [
    "Output",
    "Hand Craft?",
    "Machine",
    "Heat",
    "Power",
    "Group",
    "Tint From",
    "Skill",
    "Event",
    "Backer",
    "Wear",
    "Spark",
    "Duration",
    "Quantity",
    "XP",
    "Inputs",
]


def _get_levels(recipe):
    every = [
        recipe.output.english,
        recipe.can_hand_craft,
        recipe.machine,
        recipe.heat if recipe.machine == "FURNACE" else recipe.power,
        recipe.group_name,
        ",".join(i.english for i in recipe.tints.all()),
        ",".join(f"{r.skill.name} Lvl{r.level}" for r in recipe.requirements.all()),
        recipe.get_required_event_display(),
        recipe.required_backer_tier,
    ]

    single_row = every[:]
    bulk_row = every[:]
    mass_row = every[:]

    single_only = False

    for level in recipe.levels.all():
        if recipe.machine == "FURNACE":
            if level.level == 0:
                extra_row = [
                    level.wear,
                    level.duration,
                    level.output_quantity,
                    recipe.craft_xp * level.output_quantity,
                ]
            else:
                continue
        else:
            extra_row = [
                level.wear,
                level.spark,
                level.duration,
                level.output_quantity,
                recipe.craft_xp * level.output_quantity,
            ]

        if level.inputs.count() == 0:
            single_only = True

        for input_item in level.inputs.all():
            if input_item.group is None:
                extra_row += [input_item.item.english, input_item.count]
            else:
                extra_row += [input_item.group.name, input_item.count]

        if level.level == 0:
            single_row += extra_row
        elif level.level == 1:
            bulk_row += extra_row
        else:
            mass_row += extra_row

    return single_only, [single_row, bulk_row, mass_row]


@click.command()
def command():
    queryset = Recipe.objects.all().prefetch_related(
        "tints",
        "tints__localizedname_set",
        "requirements",
        "requirements__skill",
        "levels",
        "levels__inputs",
        "levels__inputs__group",
        "levels__inputs__item",
        "levels__inputs__item__localizedname_set",
    )
    groups = RecipeGroup.objects.all().prefetch_related(
        "members",
        "members__localizedname_set",
    )

    workbook = Workbook()
    workbook.active.title = "Single"
    workbook.create_sheet("Bulk")
    workbook.create_sheet("Mass")
    workbook.create_sheet("Furnace")
    workbook.create_sheet("Groups")

    group_sheet = workbook["Groups"]
    group_sheet.append(["Name", "Members"])

    for group in groups:
        group_sheet.append([group.name] + [i.english for i in group.members.all()])

    single = workbook["Single"]
    bulk = workbook["Bulk"]
    mass = workbook["Mass"]
    furnace = workbook["Furnace"]

    crafting_headers = HEADERS.copy()
    crafting_headers.pop(3)  # Heat
    single.append(crafting_headers)
    bulk.append(crafting_headers)
    mass.append(crafting_headers)

    furnace_headers = HEADERS.copy()
    furnace_headers.pop(11)  # Spark

    click.echo("Creating recipes...")
    with click.progressbar(queryset) as pbar:
        for recipe in pbar:
            single_only, rows = _get_levels(recipe)

            if recipe.machine == "FURNACE":
                furnace.append(rows[0])
                continue

            if rows[0][12] == rows[1][12]:
                single_only = True

            single.append(rows[0])
            if not single_only:
                bulk.append(rows[1])
                mass.append(rows[2])

    click.echo("Resizing columns...")
    set_column_widths(group_sheet)
    set_column_widths(single)
    set_column_widths(bulk)
    set_column_widths(mass)

    click.echo("Creating file...")
    create_export_file(FILENAME, "xlsx", DESCRIPTION, save_virtual_workbook(workbook))

    click.echo("Purging CDN cache...")
    purge_static_cache(["exports"])
