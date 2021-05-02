import socket
from datetime import timedelta
from email.utils import parsedate_to_datetime

import requests
from azure.identity import ClientSecretCredential
from azure.mgmt.cdn import CdnManagementClient
from azure.mgmt.cdn.models import PurgeParameters
from celery.utils.log import get_task_logger
from django.conf import settings
from django.contrib.sites.models import Site
from django.core.cache import cache
from django.db.models import Q
from django.utils import timezone
from django_celery_beat.models import IntervalSchedule, PeriodicTask
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from requests.exceptions import ReadTimeout

from boundlexx.api.utils import (
    PURGE_CACHE_LOCK,
    PURGE_CACHE_PATHS,
    PURGE_GROUPS,
    create_export_file,
    queue_purge_paths,
    set_column_widths,
)
from boundlexx.boundless.models import Item, World, WorldBlockColor
from boundlexx.boundless.utils import get_world_block_color_item_ids
from config.celery_app import app

MAX_SINGLE_PURGE = 50
WARM_CACHE_TASK = "boundlexx.api.tasks.warm_cache"
PURGE_DJ_CACHE = "boundlexx.api.tasks.purge_django_cache"
WARM_CACHE_PATHS: list[str] = [
    # "/api/v1/schema/",
    # "/api/v1/schema/?format=openapi-json",
    # "/api/v2/schema/",
    # "/api/v2/schema/?format=openapi-json",
]
STATIC_CONTAINERS: list[str] = [
    "atlas",
    "emoji",
    "exports",
    "images",
    "items",
    "logos",
    "skills",
    "worlds",
]
COLOR_EXPORT_FILENAME = "color_export"
COLOR_EXPORT_DESCRIPTION = """Export for all known block colors in the known universe.
"""

logger = get_task_logger(__name__)


def _path_chunks(iterable, chunk_size):
    while len(iterable) > chunk_size:
        yield iterable[:chunk_size]
        iterable = iterable[chunk_size:]
    yield iterable


def _reschedule_warm(run_time, path):
    _reschedule(run_time, WARM_CACHE_TASK, f"Warm Cache - {path}", f'["{path}"]')


def _reschedule(run_time, task_name, name, args="[]"):
    task = PeriodicTask.objects.filter(task=task_name, args=args).first()

    if task is None:
        interval, _ = IntervalSchedule.objects.get_or_create(
            every=1, period=IntervalSchedule.SECONDS
        )

        task = PeriodicTask.objects.create(
            task=task_name,
            name=name,
            interval=interval,
            one_off=True,
            start_time=run_time,
            enabled=True,
            args=args,
        )
    else:
        task.one_off = True
        task.start_time = run_time
        task.enabled = True
        task.save()


@app.task
def purge_django_cache():
    keys = cache.keys("views.decorators.cache.*")
    for key in keys:
        cache.delete(key)

    for path in WARM_CACHE_PATHS:
        _reschedule_warm(timezone.now(), path)


@app.task
def purge_static_cache(containers: list[str] = None):
    if (
        settings.AZURE_STATIC_CDN_ENDPOINT_NAME is None
        or len(settings.AZURE_STATIC_CDN_ENDPOINT_NAME) == 0
    ):
        logger.warning("Azure settings not configured")
        return

    if containers is None:
        containers = STATIC_CONTAINERS

    credentials = ClientSecretCredential(
        tenant_id=settings.AZURE_TENANT_ID,
        client_id=settings.AZURE_CLIENT_ID,
        client_secret=settings.AZURE_CLIENT_SECRET,
    )

    client = CdnManagementClient(credentials, settings.AZURE_SUBSCRIPTION_ID)

    paths_group: list[str] = []
    for group in containers:
        paths_group.append(f"/{settings.AZURE_CONTAINER_PREFIX}{group}/*")
    logger.info(
        "Purging paths: %s %s %s %s",
        settings.AZURE_STATIC_CDN_RESOURCE_GROUP,
        settings.AZURE_STATIC_CDN_PROFILE_NAME,
        settings.AZURE_STATIC_CDN_ENDPOINT_NAME,
        paths_group,
    )

    poller = client.endpoints.begin_purge_content(
        settings.AZURE_STATIC_CDN_RESOURCE_GROUP,
        settings.AZURE_STATIC_CDN_PROFILE_NAME,
        settings.AZURE_STATIC_CDN_ENDPOINT_NAME,
        PurgeParameters(content_paths=paths_group),
    )
    poller.result()


@app.task
def purge_cache(all_paths=False):
    # run after Azure has had a change to purge
    _reschedule(
        timezone.now() + timedelta(minutes=5),
        PURGE_DJ_CACHE,
        "Purge Django Cached Views",
    )

    if all_paths:
        paths = PURGE_GROUPS["__all__"]
    else:
        if not settings.AZURE_CDN_DYNAMIC_PURGE:
            logger.info("Dynamic purging disabled")
            return

        with cache.lock(PURGE_CACHE_LOCK, expire=10, auto_renewal=False):
            paths = cache.get(PURGE_CACHE_PATHS)

            if paths is not None:
                cache.delete(PURGE_CACHE_PATHS)

        if paths is None or len(paths) == 0:
            logger.warning("No paths to purge")
            return

        paths = list(paths)

    if (
        settings.AZURE_CDN_ENDPOINT_NAME is None
        or len(settings.AZURE_CDN_ENDPOINT_NAME) == 0
    ):
        logger.warning("Azure settings not configured")
        return

    try:
        credentials = ClientSecretCredential(
            tenant_id=settings.AZURE_TENANT_ID,
            client_id=settings.AZURE_CLIENT_ID,
            client_secret=settings.AZURE_CLIENT_SECRET,
        )

        client = CdnManagementClient(credentials, settings.AZURE_SUBSCRIPTION_ID)

        for paths_group in _path_chunks(paths, MAX_SINGLE_PURGE):
            logger.info(
                "Purging paths: %s %s %s %s",
                settings.AZURE_CDN_RESOURCE_GROUP,
                settings.AZURE_CDN_PROFILE_NAME,
                settings.AZURE_CDN_ENDPOINT_NAME,
                paths_group,
            )

            poller = client.endpoints.begin_purge_content(
                settings.AZURE_CDN_RESOURCE_GROUP,
                settings.AZURE_CDN_PROFILE_NAME,
                settings.AZURE_CDN_ENDPOINT_NAME,
                PurgeParameters(content_paths=paths_group),
            )
            poller.result()
    except (Exception, socket.gaierror) as ex:  # pylint: disable=broad-except
        logger.info("Rescheduling paths: %s", paths)
        queue_purge_paths(paths)

        if "No address associated with hostname" not in str(ex):
            raise


@app.task
def warm_cache(path):
    _warm_cache(path)


def _warm_cache(path):
    lock = cache.lock(f"boundlexx:api:warm_cache:{path}")

    acquired = lock.acquire(blocking=True, timeout=1)

    if not acquired:
        logger.info("Could not acquire lock")
        return

    rescheduled = False
    try:
        domain = Site.objects.get_current().domain

        try:
            r = requests.get(f"https://{domain}{path}", timeout=10)
        except (requests.HTTPError, ReadTimeout):
            logger.info("Timeout while making request, rescheduling...")
            _reschedule_warm(timezone.now() + timedelta(seconds=60), path)
            rescheduled = True
            return

        if not r.ok:
            logger.info("Bad response code, rescheduling...")
            _reschedule_warm(timezone.now() + timedelta(seconds=60), path)
            rescheduled = True
            return

        if "X-Cache" not in r.headers or r.headers["X-Cache"] != "HIT":
            logger.info("Response not cached, rescheduling...")
            _reschedule_warm(timezone.now() + timedelta(seconds=10), path)
            rescheduled = True
            return

        expires = parsedate_to_datetime(r.headers["Expires"])

        if expires < timezone.now():
            logger.info("Expiration date in the past! Rescheduling...")
            _reschedule_warm(timezone.now() + timedelta(minutes=5), path)
            rescheduled = True
            return

        logger.info("Request cached! Rescheduling for %s", expires)
        _reschedule_warm(expires + timedelta(seconds=5), path)
        rescheduled = True
    finally:
        if not rescheduled:
            _reschedule_warm(timezone.now() + timedelta(seconds=60), path)
        lock.release()


@app.task
def create_world_colors_export():
    workbook = Workbook()
    workbook.active.title = "Homeworlds"
    workbook.create_sheet("Exoworlds")
    workbook.create_sheet("Sovereign")

    homeworlds = workbook["Homeworlds"]
    exoworlds = workbook["Exoworlds"]
    sovereign = workbook["Sovereign"]
    sovereign.append(["Name", "Avaiable Colors"])

    worlds = World.objects.filter(owner__isnull=True, is_creative=False).order_by("id")

    logger.info("Generating Sovereign Color...")
    items = Item.objects.filter(game_id__in=get_world_block_color_item_ids()).order_by(
        "game_id"
    )
    headers = [
        "Name",
        "ID",
    ]
    for item in items:
        headers.append(item.english)

        sovereign_columns = [item.english]
        wbcs = (
            WorldBlockColor.objects.filter(item=item, is_default=True)
            .filter(
                Q(world__isnull=True)
                | Q(world__end__isnull=True, world__is_creative=False)
                | Q(world__owner__isnull=False, world__is_creative=False)
            )
            .select_related("item")
            .order_by("color__game_id")
            .distinct("color__game_id")
        )

        for wbc in wbcs:
            sovereign_columns.append(f"{wbc.color.default_name} ({wbc.color.game_id})")

        sovereign.append(sovereign_columns)

    logger.info("Generating World Rows...")
    homeworlds.append(headers)
    exoworlds.append(headers)
    for world in worlds:
        columns = [world.display_name, world.id]

        for wbc in world.worldblockcolor_set.filter(is_default=True).order_by(
            "item__game_id"
        ):
            columns.append(f"{wbc.color.default_name} ({wbc.color.game_id})")

        if world.is_exo:
            exoworlds.append(columns)
        else:
            homeworlds.append(columns)

    set_column_widths(homeworlds)
    set_column_widths(exoworlds)
    set_column_widths(sovereign)

    create_export_file(
        COLOR_EXPORT_FILENAME,
        "xlsx",
        COLOR_EXPORT_DESCRIPTION,
        save_virtual_workbook(workbook),
    )
